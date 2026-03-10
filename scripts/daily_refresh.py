#!/usr/bin/env python3
import argparse
import csv
import io
import json
import math
import os
import ssl
import sys
import time
from math import ceil
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen
import urllib.error

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = Path("/etc/macro-monitor.env")
if ENV_FILE.exists():
    try:
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            key = k.strip()
            val = v.strip()
            if key and val and not os.environ.get(key):
                os.environ[key] = val
    except Exception:
        pass
sys.path.insert(0, str(ROOT / "server"))
from db import (
    get_api_key,
    init_db,
    log_token_usage,
    replace_sheet_rows,
    save_model_snapshot,
    upsert_daily_report,
    upsert_daily_report_ai_insight,
)

MODEL_PATH = ROOT / "model.xlsx"
REPORTS_DIR = ROOT / "reports"
DATA_DIR = ROOT / "data"
CTX = ssl._create_unverified_context()
UPDATE_LOG_PATH = ROOT / "data_update_log.json"
FRED_API_BASE = "https://api.stlouisfed.org/fred/series/observations"
_FRED_KEY_CACHE = None
_OPENAI_KEY_CACHE = None
OPENAI_MODEL = str(os.environ.get("OPENAI_MODEL", "gpt-5.4")).strip()
OPENAI_BASE_URL = str(os.environ.get("OPENAI_BASE_URL", "https://api.openai.com/v1")).strip().rstrip("/")
OPENAI_MAX_RETRIES = max(1, int(os.environ.get("OPENAI_MAX_RETRIES", "2")))
OPENAI_STRICT_ANALYSIS = str(os.environ.get("OPENAI_STRICT_ANALYSIS", "false")).strip().lower() in {"1", "true", "yes", "on"}
AI_PROMPT_VERSION = "daily_report_v1"


def fetch_text(url: str) -> str:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30, context=CTX) as resp:
        return resp.read().decode("utf-8", errors="ignore")


def fetch_json(url: str):
    return json.loads(fetch_text(url))


def post_json(url: str, payload: dict, headers=None):
    req = Request(url, method="POST")
    req.add_header("Content-Type", "application/json")
    for k, v in (headers or {}).items():
        req.add_header(k, v)
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    with urlopen(req, body, timeout=60, context=CTX) as resp:
        return json.loads(resp.read().decode("utf-8", errors="ignore"))


def get_fred_api_key():
    global _FRED_KEY_CACHE
    if _FRED_KEY_CACHE is not None:
        return _FRED_KEY_CACHE
    key = str(os.environ.get("FRED_API_KEY", "")).strip()
    if key:
        _FRED_KEY_CACHE = key
        return _FRED_KEY_CACHE
    try:
        key = str(get_api_key("fred") or "").strip()
    except Exception:
        key = ""
    _FRED_KEY_CACHE = key
    return _FRED_KEY_CACHE


def get_openai_api_key():
    global _OPENAI_KEY_CACHE
    if _OPENAI_KEY_CACHE is not None:
        return _OPENAI_KEY_CACHE
    key = str(os.environ.get("OPENAI_API_KEY", "")).strip()
    if key:
        _OPENAI_KEY_CACHE = key
        return _OPENAI_KEY_CACHE
    try:
        key = str(get_api_key("openai") or "").strip()
    except Exception:
        key = ""
    _OPENAI_KEY_CACHE = key
    return _OPENAI_KEY_CACHE


def _extract_chat_text(resp: dict):
    if not isinstance(resp, dict):
        return ""
    choices = resp.get("choices") or []
    if not choices:
        return ""
    msg = choices[0].get("message") or {}
    return str(msg.get("content") or "").strip()


def build_local_ai_insight(today, total_score, model_status, top_dimension_contributors, failed_count, short_summary):
    top = [str(x.get("name") or x.get("id") or "--") for x in (top_dimension_contributors or [])[:3]]
    short_zh = f"{today} 总分 {total_score}（{model_status}），重点关注 {', '.join(top) if top else '关键维度'}。"
    short_en = f"{today} score {total_score} ({model_status}); focus on {', '.join(top) if top else 'key dimensions'}."
    detailed_zh = (
        f"## 总览\n- 日期: {today}\n- 综合得分: {total_score}\n- 状态: {model_status}\n\n"
        f"## 结构解读\n- 主要贡献维度: {', '.join(top) if top else '--'}\n- 数据更新失败指标数: {failed_count}\n\n"
        f"## 结论\n- {short_summary}"
    )
    detailed_en = (
        f"## Overview\n- Date: {today}\n- Composite score: {total_score}\n- Regime: {model_status}\n\n"
        f"## Structure\n- Top contributing dimensions: {', '.join(top) if top else '--'}\n- Failed indicators in online refresh: {failed_count}\n\n"
        f"## Conclusion\n- {short_summary}"
    )
    return {
        "short_summary_zh": short_zh,
        "short_summary_en": short_en,
        "detailed_markdown_zh": detailed_zh,
        "detailed_markdown_en": detailed_en,
        "key_risks": [],
        "action_items": [],
    }, "local-fallback", ""


def generate_ai_insight(
    today,
    generated_at,
    total_score,
    model_status,
    top_dimension_contributors,
    trigger_alerts,
    key_indicators_snapshot,
    indicator_details,
    short_summary,
):
    api_key = get_openai_api_key()
    if not api_key:
        fallback, model_name, _ = build_local_ai_insight(
            today, total_score, model_status, top_dimension_contributors, len(indicator_details or []), short_summary
        )
        return fallback, model_name, "openai_missing_api_key"

    compact_failed = [
        {
            "code": x.get("IndicatorCode"),
            "name": x.get("IndicatorName"),
            "valueDate": x.get("ValueDate"),
            "status": x.get("VerificationStatus"),
            "error": x.get("VerificationError"),
        }
        for x in (indicator_details or [])
        if not bool(x.get("VerifiedOnline"))
    ][:15]
    compact_context = {
        "date": today,
        "generatedAt": generated_at,
        "totalScore": total_score,
        "status": model_status,
        "topDimensions": (top_dimension_contributors or [])[:6],
        "triggeredAlerts": [x for x in (trigger_alerts or []) if x.get("triggered")][:10],
        "keyIndicatorsSnapshot": (key_indicators_snapshot or [])[:10],
        "failedIndicators": compact_failed,
        "baseSummary": short_summary,
    }
    system_text = (
        "You are a macro risk analyst. Return STRICT JSON only with keys: "
        "short_summary_zh, short_summary_en, detailed_markdown_zh, detailed_markdown_en, key_risks, action_items. "
        "short_summary fields <= 120 Chinese chars / <= 220 English chars. "
        "Detailed markdown should be concise and data-grounded. Do not fabricate unavailable values."
    )
    user_text = f"Context JSON:\n{json.dumps(compact_context, ensure_ascii=False)}"
    body = {
        "model": OPENAI_MODEL,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": system_text},
            {"role": "user", "content": user_text},
        ],
    }

    last_err = ""
    for i in range(OPENAI_MAX_RETRIES):
        try:
            resp = post_json(
                f"{OPENAI_BASE_URL}/chat/completions",
                body,
                headers={"Authorization": f"Bearer {api_key}"},
            )
            txt = _extract_chat_text(resp)
            parsed = json.loads(txt)
            if isinstance(parsed, dict) and parsed.get("short_summary_zh") and parsed.get("detailed_markdown_zh"):
                return parsed, OPENAI_MODEL, ""
            last_err = "invalid_json_shape"
        except urllib.error.HTTPError as he:
            detail = ""
            try:
                detail = he.read().decode("utf-8", errors="ignore")[:320]
            except Exception:
                detail = ""
            last_err = f"http_{he.code}:{detail}" if detail else f"http_{he.code}"
            if he.code == 429 and i < OPENAI_MAX_RETRIES - 1:
                time.sleep(1.2 * (i + 1))
                continue
        except Exception as e:
            last_err = str(e)[:180]
        if i < OPENAI_MAX_RETRIES - 1:
            time.sleep(0.8 * (i + 1))

    fallback, model_name, _ = build_local_ai_insight(
        today, total_score, model_status, top_dimension_contributors, len(compact_failed), short_summary
    )
    return fallback, model_name, f"openai_failed:{last_err}"


def fred_observations(series: str):
    key = get_fred_api_key()
    if not key:
        return []
    query = urlencode(
        {
            "series_id": series,
            "api_key": key,
            "file_type": "json",
            "sort_order": "asc",
            "observation_start": "1970-01-01",
        }
    )
    data = fetch_json(f"{FRED_API_BASE}?{query}")
    out = []
    for row in data.get("observations") or []:
        d = str(row.get("date") or "")
        v = str(row.get("value") or "")
        if d and v and v != ".":
            out.append((d, float(v)))
    return out


def fred_last(series: str):
    vals = fred_observations(series)
    if vals:
        return vals[-1]
    txt = fetch_text(f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}")
    rows = list(csv.reader(io.StringIO(txt)))
    for d, v in reversed(rows[1:]):
        if v and v != ".":
            return d, float(v)
    raise ValueError(f"No value for {series}")


def fred_yoy(series: str):
    vals = fred_observations(series)
    if not vals:
        txt = fetch_text(f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}")
        rows = list(csv.reader(io.StringIO(txt)))[1:]
        vals = [(d, float(v)) for d, v in rows if v and v != "."]
    if len(vals) < 13:
        raise ValueError(f"Insufficient history for {series}")
    d, v = vals[-1]
    _, v12 = vals[-13]
    return d, (v / v12 - 1) * 100


def first_series_token(raw: str) -> str:
    text = str(raw or "").upper()
    for token in text.replace("&", " ").replace("/", " ").replace(",", " ").split():
        token = token.strip()
        if token and token[0].isalpha() and all(ch.isalnum() or ch == "_" for ch in token):
            return token
    return ""


def clamp(v, lo, hi):
    return max(lo, min(hi, v))


def as_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def serializable(v):
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    return str(v)


def sheet_to_dicts(ws):
    header = None
    rows = []
    for r in ws.iter_rows(values_only=True):
        vals = [serializable(x) for x in r]
        if header is None:
            if vals and vals[0] and any(vals[1:]):
                header = vals
            continue
        if not any(v is not None and str(v).strip() != "" for v in vals):
            continue
        row = {}
        for i, h in enumerate(header):
            if h is None:
                continue
            if i < len(vals):
                row[str(h)] = vals[i]
        rows.append(row)
    return rows


def status_from_score(score):
    if score >= 75:
        return "扩张偏热"
    if score >= 60:
        return "温和扩张"
    if score >= 45:
        return "中性脆弱"
    if score >= 30:
        return "防御区"
    return "衰退/危机"


def parse_date_safe(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Handle common formats found in workbook/API payloads.
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"):
        try:
            d = datetime.strptime(s, fmt).date()
            if fmt in ("%Y-%m", "%Y/%m"):
                d = d.replace(day=1)
            return d
        except Exception:
            continue
    try:
        # ISO datetime support.
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except Exception:
        return None


def max_age_days_by_frequency(raw_freq):
    f = str(raw_freq or "").strip().lower()
    if "day" in f or "daily" in f:
        return 3
    if "week" in f or "weekly" in f:
        return 10
    if "month" in f or "monthly" in f:
        return 45
    if "quarter" in f or "quarterly" in f:
        return 120
    if "year" in f or "annual" in f:
        return 400
    # Unknown cadence defaults to monthly tolerance.
    return 45


def is_online_fetchable(indicator):
    code = str(indicator.get("IndicatorCode") or "")
    source = str(indicator.get("Source") or "").lower()
    series_hint = indicator.get("Series")
    if code in {
        "YC_10Y3M",
        "SOFR",
        "FED_ASSETS",
        "NET_LIQ_PROXY",
        "GDP_QOQ_SAAR",
        "CLAIMS_4WMA",
        "CORE_CPI_YOY",
        "CORE_PCE_YOY",
        "BREAKEVEN_5Y5Y",
        "UNRATE",
        "WAGE_GROWTH",
        "CC_DELINQ",
        "REAL_DISP_INC",
        "HY_OAS",
        "MORTGAGE_30Y",
        "HOUSING_STARTS",
        "VIX",
        "DXY",
        "US_JP_10Y_SPREAD",
        "FCI",
        "BANK_LENDING",
        "TED_SPREAD",
        "WTI",
        "BTC",
        "USDC_MCAP",
    }:
        return True
    if "fred" in source and first_series_token(series_hint):
        return True
    return False


def load_last_online_update():
    if not UPDATE_LOG_PATH.exists():
        return {"updated_count": 0, "failed_count": 0, "updated": [], "failed": []}
    try:
        payload = json.loads(UPDATE_LOG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"updated_count": 0, "failed_count": 0, "updated": [], "failed": []}
    if not isinstance(payload, dict):
        return {"updated_count": 0, "failed_count": 0, "updated": [], "failed": []}
    return {
        "updated_count": int(payload.get("updated_count") or 0),
        "failed_count": int(payload.get("failed_count") or 0),
        "updated": payload.get("updated") or [],
        "failed": payload.get("failed") or [],
    }


def run(mode="full", report_date=None, strict_freshness=False, require_openai_ai=False):
    if mode not in {"full", "fetch-only", "report-only"}:
        raise ValueError(f"invalid mode: {mode}")
    today = str(report_date or date.today().isoformat())
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    fetch_enabled = mode in {"full", "fetch-only"}
    report_enabled = mode in {"full", "report-only"}

    wb = load_workbook(MODEL_PATH)
    ws_dim = wb["Dimensions"]
    ws_ind = wb["Indicators"]
    ws_in = wb["Inputs"]

    indicators = {}
    for r in range(2, ws_ind.max_row + 1):
        code = ws_ind.cell(r, 1).value
        if not code:
            continue
        indicators[str(code)] = {
            "IndicatorCode": str(code),
            "row": r,
            "DimensionID": ws_ind.cell(r, 2).value,
            "IndicatorName": ws_ind.cell(r, 3).value,
            "Frequency": ws_ind.cell(r, 5).value,
            "Source": ws_ind.cell(r, 6).value,
            "SourceURL": ws_ind.cell(r, 7).value,
            "Series": ws_ind.cell(r, 8).value,
            "ScaleType": ws_ind.cell(r, 9).value,
            "Direction": ws_ind.cell(r, 10).value,
            "Best": ws_ind.cell(r, 11).value,
            "Worst": ws_ind.cell(r, 12).value,
            "WorstLow": ws_ind.cell(r, 13).value,
            "TargetLow": ws_ind.cell(r, 14).value,
            "TargetHigh": ws_ind.cell(r, 15).value,
            "WorstHigh": ws_ind.cell(r, 16).value,
            "CapLow": ws_ind.cell(r, 17).value,
            "CapHigh": ws_ind.cell(r, 18).value,
            "WeightWithinDim": ws_ind.cell(r, 19).value,
        }

    input_rows = {}
    for r in range(7, ws_in.max_row + 1):
        code = ws_in.cell(r, 1).value
        if code:
            input_rows[str(code)] = r

    ws_in["B2"] = today
    cache = {}

    def s_last(series):
        if series in cache:
            return cache[series]
        cache[series] = fred_last(series)
        return cache[series]

    updated = []
    failed = []

    if fetch_enabled:
        for code in indicators:
            try:
                source = str(indicators[code].get("Source") or "").lower()
                series_hint = indicators[code].get("Series")
                d = None
                v = None
                if code == "YC_10Y3M":
                    d1, x = s_last("DGS10")
                    d2, y = s_last("DGS3MO")
                    d = max(d1, d2)
                    v = (x - y) * 100
                elif code == "SOFR":
                    d, v = s_last("SOFR")
                elif code == "FED_ASSETS":
                    d, x = s_last("WALCL")
                    v = x / 1_000_000
                elif code == "NET_LIQ_PROXY":
                    d1, a = s_last("WALCL")
                    d2, tga = s_last("WTREGEN")
                    d3, rrp = s_last("RRPONTSYD")
                    d = max(d1, d2, d3)
                    v = (a / 1000 - tga - rrp) / 1000
                elif code == "GDP_QOQ_SAAR":
                    d, v = s_last("A191RL1Q225SBEA")
                elif code == "CLAIMS_4WMA":
                    d, v = s_last("IC4WSA")
                elif code == "CORE_CPI_YOY":
                    d, v = fred_yoy("CPILFESL")
                elif code == "CORE_PCE_YOY":
                    d, v = fred_yoy("PCEPILFE")
                elif code == "BREAKEVEN_5Y5Y":
                    d, v = s_last("T5YIFR")
                elif code == "UNRATE":
                    d, v = s_last("UNRATE")
                elif code == "WAGE_GROWTH":
                    d, v = fred_yoy("CES0500000003")
                elif code == "CC_DELINQ":
                    d, v = s_last("DRCCLACBS")
                elif code == "REAL_DISP_INC":
                    d, v = fred_yoy("DSPIC96")
                elif code == "HY_OAS":
                    d, x = s_last("BAMLH0A0HYM2")
                    v = x * 100 if x < 50 else x
                elif code == "MORTGAGE_30Y":
                    d, v = s_last("MORTGAGE30US")
                elif code == "HOUSING_STARTS":
                    d, x = s_last("HOUST")
                    v = x / 1000
                elif code == "VIX":
                    d, v = s_last("VIXCLS")
                elif code == "DXY":
                    d, v = s_last("DTWEXBGS")
                elif code == "US_JP_10Y_SPREAD":
                    d1, us10 = s_last("DGS10")
                    d2, jp10 = s_last("IRLTLT01JPM156N")
                    d = max(d1, d2)
                    v = (us10 - jp10) * 100
                elif code == "FCI":
                    d, v = s_last("NFCI")
                elif code == "BANK_LENDING":
                    d, v = fred_yoy("TOTBKCR")
                elif code == "TED_SPREAD":
                    d, v = s_last("TEDRATE")
                elif code == "WTI":
                    d, v = s_last("DCOILWTICO")
                elif code == "BTC":
                    j = fetch_json("https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd")
                    d = today
                    v = float(j["bitcoin"]["usd"])
                elif code == "USDC_MCAP":
                    j = fetch_json("https://api.coingecko.com/api/v3/coins/usd-coin")
                    d = today
                    v = float(j["market_data"]["market_cap"]["usd"]) / 1_000_000_000
                elif "fred" in source:
                    token = first_series_token(series_hint)
                    if not token:
                        raise ValueError("no FRED series token")
                    d, v = s_last(token)
                else:
                    raise ValueError("manual/proprietary source required")

                row = input_rows.get(code)
                if row and v is not None:
                    ws_in.cell(row, 2).value = round(v, 4)
                    ws_in.cell(row, 3).value = d
                    ws_in.cell(row, 4).value = today
                updated.append({"code": code, "value": round(v, 6), "date": d})
            except Exception as e:
                failed.append({"code": code, "error": str(e)[:180]})

    dimensions = []
    for r in range(2, ws_dim.max_row + 1):
        did = ws_dim.cell(r, 1).value
        if not did:
            continue
        dimensions.append({
            "id": str(did),
            "name": ws_dim.cell(r, 2).value,
            "tier": ws_dim.cell(r, 3).value,
            "weight": as_number(ws_dim.cell(r, 4).value) or 0,
            "definition": ws_dim.cell(r, 5).value,
            "update": ws_dim.cell(r, 6).value,
        })

    input_values = {}
    input_meta = {}
    for code, r in input_rows.items():
        v = as_number(ws_in.cell(r, 2).value)
        value_date = serializable(ws_in.cell(r, 3).value)
        source_date = serializable(ws_in.cell(r, 4).value)
        input_meta[code] = {"value_date": value_date, "source_date": source_date}
        if v is not None:
            input_values[code] = v

    indicator_scores = []
    for code, m in indicators.items():
        if code not in input_values:
            continue
        x = input_values[code]
        cap_low = as_number(m["CapLow"])
        cap_high = as_number(m["CapHigh"])
        if cap_low is not None:
            x = max(x, cap_low)
        if cap_high is not None:
            x = min(x, cap_high)

        scale = str(m["ScaleType"] or "").lower()
        direction = str(m["Direction"] or "").lower()
        best = as_number(m["Best"])
        worst = as_number(m["Worst"])
        wl = as_number(m["WorstLow"])
        tl = as_number(m["TargetLow"])
        th = as_number(m["TargetHigh"])
        wh = as_number(m["WorstHigh"])
        w = as_number(m["WeightWithinDim"]) or 0

        score = None
        if "targetband" in scale and None not in (wl, tl, th, wh):
            if tl <= x <= th:
                score = 100
            elif x < tl:
                score = ((x - wl) / (tl - wl)) * 100
            else:
                score = ((wh - x) / (wh - th)) * 100
        elif "higher" in direction and None not in (best, worst) and best != worst:
            score = ((x - worst) / (best - worst)) * 100
        elif "lower" in direction and None not in (best, worst) and worst != best:
            score = ((worst - x) / (worst - best)) * 100

        if score is None or not math.isfinite(score):
            continue
        score = clamp(score, 0, 100)
        indicator_scores.append({
            "IndicatorCode": code,
            "DimensionID": m["DimensionID"],
            "IndicatorName": m["IndicatorName"],
            "LatestValue": round(input_values[code], 4),
            "Score(0-100)": round(score, 2),
            "WeightWithinDim": round(w, 4),
            "WeightedScore": round(score * w, 4),
        })

    dim_summary = []
    for d in dimensions:
        rows = [r for r in indicator_scores if str(r["DimensionID"]) == d["id"]]
        wsum = sum(r["WeightWithinDim"] for r in rows)
        wscore = sum(r["WeightedScore"] for r in rows)
        dim_score = wscore / wsum if wsum > 0 else 0
        contrib = dim_score * d["weight"] / 100
        dim_summary.append({
            "id": d["id"],
            "name": d["name"],
            "tier": d["tier"],
            "weight": d["weight"],
            "score": round(dim_score, 2),
            "contribution": round(contrib, 4),
            "definition": d["definition"],
            "update": d["update"],
        })

    total_score = round(sum(x["contribution"] for x in dim_summary), 2)
    model_status = status_from_score(total_score)

    val = input_values
    alerts = [
        {"id": "A01", "level": "RED", "condition": "VIX > 30", "triggered": (val.get("VIX", -999) > 30)},
        {"id": "A02", "level": "RED", "condition": "MOVE > 140", "triggered": (val.get("MOVE", -999) > 140)},
        {"id": "A03", "level": "YELLOW", "condition": "HY OAS > 600bps", "triggered": (val.get("HY_OAS", -999) > 600)},
        {"id": "A04", "level": "YELLOW", "condition": "10Y-3M < -50bps", "triggered": (val.get("YC_10Y3M", 999) < -50)},
        {"id": "A05", "level": "YELLOW", "condition": "失业率 > 6%", "triggered": (val.get("UNRATE", -999) > 6)},
        {"id": "A06", "level": "YELLOW", "condition": "核心PCE > 3.5%", "triggered": (val.get("CORE_PCE_YOY", -999) > 3.5)},
        {"id": "A07", "level": "YELLOW", "condition": "WTI > 100", "triggered": (val.get("WTI", -999) > 100)},
    ]

    ranked = sorted(dim_summary, key=lambda x: x["score"])
    weak = ranked[:3]
    strong = list(reversed(ranked[-3:]))

    drivers = [
        {"title": "Primary Support", "text": f"{strong[0]['name']} is the strongest block ({strong[0]['score']})." if strong else ""},
        {"title": "Primary Drag", "text": f"{weak[0]['name']} is the key drag ({weak[0]['score']})." if weak else ""},
        {"title": "Risk Trigger", "text": f"{sum(1 for a in alerts if a['triggered'])} alert(s) currently triggered."},
    ]

    key_watch = [
        {"label": "YC_10Y3M", "value": val.get("YC_10Y3M")},
        {"label": "VIX", "value": val.get("VIX")},
        {"label": "HY_OAS", "value": val.get("HY_OAS")},
        {"label": "CORE_CPI_YOY", "value": val.get("CORE_CPI_YOY")},
        {"label": "UNRATE", "value": val.get("UNRATE")},
        {"label": "WTI", "value": val.get("WTI")},
        {"label": "BTC", "value": val.get("BTC")},
    ]
    key_indicators_snapshot = [
        {"title": "10Y-3M利差", "label": "YC_10Y3M", "value": val.get("YC_10Y3M"), "source": "FRED"},
        {"title": "VIX指数", "label": "VIX", "value": val.get("VIX"), "source": "FRED/CBOE"},
        {"title": "HY OAS", "label": "HY_OAS", "value": val.get("HY_OAS"), "source": "FRED"},
        {"title": "核心CPI同比", "label": "CORE_CPI_YOY", "value": val.get("CORE_CPI_YOY"), "source": "FRED/BLS"},
        {"title": "失业率", "label": "UNRATE", "value": val.get("UNRATE"), "source": "FRED/BLS"},
        {"title": "WTI油价", "label": "WTI", "value": val.get("WTI"), "source": "FRED"},
    ]
    top_dimension_contributors = sorted(
        [{"name": d["name"], "score": d["score"], "contribution": d["contribution"], "id": d["id"]} for d in dim_summary],
        key=lambda x: x["contribution"],
        reverse=True,
    )
    daily_watched_items = [f"{x['label']}: {x['value']}" for x in key_watch if x.get("value") is not None]

    updated_set = {x["code"] for x in updated}
    failed_map = {x["code"]: x.get("error", "") for x in failed}
    today_date = parse_date_safe(today)
    freshness_checks = []
    freshness_failures = []
    online_fetchable_total = 0
    online_verified_count = 0

    for code, m in indicators.items():
        online_fetchable = is_online_fetchable(m)
        if online_fetchable:
            online_fetchable_total += 1
        verified = code in updated_set
        if verified:
            online_verified_count += 1
        value_date = input_meta.get(code, {}).get("value_date")
        parsed_value_date = parse_date_safe(value_date)
        max_age_days = max_age_days_by_frequency(m.get("Frequency"))
        age_days = None
        if parsed_value_date and today_date:
            age_days = max(0, (today_date - parsed_value_date).days)
        stale = age_days is None or age_days > max_age_days
        must_pass = bool(online_fetchable)
        passed = (verified and not stale) if must_pass else True
        item = {
            "IndicatorCode": code,
            "Frequency": m.get("Frequency"),
            "OnlineFetchable": online_fetchable,
            "VerifiedOnline": verified,
            "ValueDate": value_date,
            "AgeDays": age_days,
            "MaxAllowedAgeDays": max_age_days,
            "Passed": passed,
            "Error": failed_map.get(code, ""),
        }
        freshness_checks.append(item)
        if must_pass and not passed:
            freshness_failures.append(item)

    indicator_details = []
    for code, m in indicators.items():
        verified = code in updated_set
        verify_status = "Verified Online" if verified else "Fallback (latest available in model inputs)"
        indicator_details.append(
            {
                "IndicatorCode": code,
                "IndicatorName": m["IndicatorName"],
                "DimensionID": m["DimensionID"],
                "Source": m["Source"],
                "Series/Code": m["Series"],
                "LatestValue": input_values.get(code),
                "ValueDate": input_meta.get(code, {}).get("value_date"),
                "SourceDate": input_meta.get(code, {}).get("source_date"),
                "VerifiedOnline": verified,
                "VerificationStatus": verify_status,
                "VerificationError": failed_map.get(code, ""),
                "FreshnessAgeDays": next((x.get("AgeDays") for x in freshness_checks if x["IndicatorCode"] == code), None),
                "FreshnessMaxAgeDays": next((x.get("MaxAllowedAgeDays") for x in freshness_checks if x["IndicatorCode"] == code), None),
                "FreshnessPassed": next((x.get("Passed") for x in freshness_checks if x["IndicatorCode"] == code), None),
                "GeneratedAt": generated_at,
            }
        )

    all14_dimensions_detailed = [
        {
            "DimensionID": d["id"],
            "DimensionName": d["name"],
            "Tier": d["tier"],
            "Weight(%)": d["weight"],
            "Definition": d["definition"],
            "Typical Update": d["update"],
            "DimScore": d["score"],
            "WeightedContribution": d["contribution"],
        }
        for d in dim_summary
    ]

    wb.save(MODEL_PATH)
    REPORTS_DIR.mkdir(exist_ok=True)
    DATA_DIR.mkdir(exist_ok=True)

    summary_updated_count = len(updated)
    summary_failed_count = len(failed)
    if not fetch_enabled:
        last_update = load_last_online_update()
        if last_update["updated_count"] or last_update["failed_count"]:
            summary_updated_count = last_update["updated_count"]
            summary_failed_count = last_update["failed_count"]

    short_summary = (
        f"Macro model updated on {today}. Public-source updater refreshed {summary_updated_count} indicators; "
        f"{summary_failed_count} indicators still require manual/proprietary updates. "
        f"Online freshness pass: {online_verified_count}/{online_fetchable_total}. "
        f"Composite score: {total_score} ({model_status})."
    )

    ai_payload, ai_model, ai_error = generate_ai_insight(
        today=today,
        generated_at=generated_at,
        total_score=total_score,
        model_status=model_status,
        top_dimension_contributors=top_dimension_contributors,
        trigger_alerts=alerts,
        key_indicators_snapshot=key_indicators_snapshot,
        indicator_details=indicator_details,
        short_summary=short_summary,
    )
    ai_short_zh = str(ai_payload.get("short_summary_zh") or "").strip()
    ai_short_en = str(ai_payload.get("short_summary_en") or "").strip()
    ai_detailed_zh = str(ai_payload.get("detailed_markdown_zh") or "").strip()
    ai_detailed_en = str(ai_payload.get("detailed_markdown_en") or "").strip()
    require_openai_effective = bool(require_openai_ai or OPENAI_STRICT_ANALYSIS)
    if report_enabled and require_openai_effective and (ai_model != OPENAI_MODEL or ai_error):
        raise RuntimeError(f"openai_analysis_required_but_failed model={ai_model} error={ai_error}")

    if strict_freshness and freshness_failures:
        preview = ", ".join(
            [
                f"{x['IndicatorCode']}[verified={x['VerifiedOnline']},age={x['AgeDays']},max={x['MaxAllowedAgeDays']}]"
                for x in freshness_failures[:12]
            ]
        )
        raise RuntimeError(
            f"strict_freshness_failed count={len(freshness_failures)} online_verified={online_verified_count}/{online_fetchable_total} failures={preview}"
        )

    report_text = ""
    if report_enabled:
        report_text_lines = [
            f"Macro Daily Report ({today})",
            f"Model As-Of: {today}",
            f"Composite Score: {total_score} ({model_status})",
            "",
            "Key Indicators To Watch",
        ]
        for item in key_watch:
            if item["value"] is not None:
                report_text_lines.append(f"- {item['label']}: {round(item['value'],4)}")
        report_text_lines += [
            "",
            "Short Summary",
            f"- {short_summary}",
            f"- Data generated at: {generated_at}",
            "- Weakest dimensions: " + ", ".join([f"{x['id']} {x['name']} ({x['score']})" for x in weak]),
            "- Strongest dimensions: " + ", ".join([f"{x['id']} {x['name']} ({x['score']})" for x in strong]),
        ]
        report_text = "\n".join(report_text_lines) + "\n"
        (REPORTS_DIR / f"{today}.txt").write_text(report_text, encoding="utf-8")

        def render_dim_cards(items):
            parts = []
            for d in items:
                related = [x for x in indicator_scores if str(x["DimensionID"]) == d["id"]][:3]
                li = "".join([f"<li><strong>{r['IndicatorCode']}</strong>: {r['LatestValue']} | score {r['Score(0-100)']}</li>" for r in related])
                parts.append(f"<div class='preview-dim-card'><strong>{d['id']} {d['name']}</strong><div>Weight {d['weight']}% | Score {d['score']} | Contribution {d['contribution']}</div><ul>{li}</ul></div>")
            return "".join(parts)

        tiers = {}
        for d in dim_summary:
            tiers.setdefault(d["tier"] or "Other", []).append(d)
        tier_html = ""
        for tier, items in tiers.items():
            tier_html += f"<section class='preview-tier'><h2>{tier}</h2>{render_dim_cards(items)}</section>"

        report_html = f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Macro Daily Report {today}</title><link rel='stylesheet' href='../styles.css'></head><body><main class='page'><section class='panel preview-header'><h1>Macro Daily Report ({today})</h1><p>Composite Score: {total_score} ({model_status})</p><p>{short_summary}</p></section><section class='panel preview-section'><h2>Key Indicators To Watch</h2><ul class='preview-list'>{''.join([f"<li>{i['label']}: {round(i['value'],4)}</li>" for i in key_watch if i['value'] is not None])}</ul></section>{tier_html}</main></body></html>"""
        (REPORTS_DIR / f"{today}.html").write_text(report_html, encoding="utf-8")

    ws_daily = wb["DailyReports"] if "DailyReports" in wb.sheetnames else wb.create_sheet("DailyReports")
    if ws_daily.max_row < 1 or ws_daily.cell(1, 1).value != "Date":
        ws_daily.cell(1, 1).value = "Date"
        ws_daily.cell(1, 2).value = "AsOf"
        ws_daily.cell(1, 3).value = "TotalScore"
        ws_daily.cell(1, 4).value = "Status"
        ws_daily.cell(1, 5).value = "Summary"
        ws_daily.cell(1, 6).value = "GeneratedAt"
        ws_daily.cell(1, 7).value = "ReportPath"

    existing_row = None
    for rr in range(2, ws_daily.max_row + 1):
        if str(ws_daily.cell(rr, 1).value or "") == today:
            existing_row = rr
            break
    target_row = existing_row or (ws_daily.max_row + 1)
    ws_daily.cell(target_row, 1).value = today
    ws_daily.cell(target_row, 2).value = today
    ws_daily.cell(target_row, 3).value = total_score
    ws_daily.cell(target_row, 4).value = model_status
    ws_daily.cell(target_row, 5).value = short_summary
    ws_daily.cell(target_row, 6).value = generated_at
    ws_daily.cell(target_row, 7).value = f"reports/{today}.html"

    index_path = REPORTS_DIR / "index.json"
    reports = []
    if index_path.exists():
        try:
            reports = json.loads(index_path.read_text(encoding="utf-8")).get("reports", [])
        except Exception:
            reports = []

    today_entry = {
        "date": today,
        "meta": {"score": str(total_score), "status": model_status},
        "text": report_text,
        "path": f"reports/{today}.html",
        "reportPayload": {
            "topDimensionContributors": top_dimension_contributors,
            "triggerAlerts": alerts,
            "dailyWatchedItems": daily_watched_items,
            "primaryDrivers": drivers,
            "keyIndicatorsSnapshot": key_indicators_snapshot,
            "all14DimensionsDetailed": all14_dimensions_detailed,
            "latestReportSummary": ai_short_zh or short_summary,
            "indicatorDetails": indicator_details,
            "aiInsight": {
                "short_summary_zh": ai_short_zh,
                "short_summary_en": ai_short_en,
                "detailed_markdown_zh": ai_detailed_zh,
                "detailed_markdown_en": ai_detailed_en,
                "model": ai_model,
                "prompt_version": AI_PROMPT_VERSION,
                "generated_at": generated_at,
                "error": ai_error,
            },
            "generatedAt": generated_at,
        },
    }
    if report_enabled:
        merged = [today_entry] + [r for r in reports if r.get("date") != today]
        index_path.write_text(json.dumps({"reports": merged}, ensure_ascii=False, indent=2), encoding="utf-8")

    snapshot = {
        "asOf": today,
        "reportDate": today if report_enabled else "",
        "totalScore": total_score,
        "status": model_status,
        "alerts": alerts,
        "dimensions": [{"name": d["name"], "score": d["score"], "contribution": d["contribution"], "id": d["id"]} for d in dim_summary],
        "drivers": drivers,
        "keyWatch": key_watch,
        "latestReportSummary": ai_short_zh or short_summary,
        "generatedAt": generated_at,
        "topDimensionContributors": top_dimension_contributors,
        "triggerAlerts": alerts,
        "dailyWatchedItems": daily_watched_items,
        "primaryDrivers": drivers,
        "keyIndicatorsSnapshot": key_indicators_snapshot,
        "all14DimensionsDetailed": all14_dimensions_detailed,
        "indicatorDetails": indicator_details,
        "tables": {
            "dimensions": sheet_to_dicts(ws_dim),
            "indicators": sheet_to_dicts(ws_ind),
            "inputs": sheet_to_dicts(ws_in),
            "scores": indicator_scores,
            "alerts": alerts,
        },
        "onlineUpdate": {"updated_count": len(updated), "failed_count": len(failed), "updated": updated, "failed": failed},
        "freshness": {
            "strict": bool(strict_freshness),
            "online_fetchable_total": online_fetchable_total,
            "online_verified_count": online_verified_count,
            "failed_count": len(freshness_failures),
            "checks": freshness_checks,
        },
    }
    (DATA_DIR / "latest_snapshot.json").write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    (ROOT / "data_update_log.json").write_text(json.dumps(snapshot["onlineUpdate"], ensure_ascii=False, indent=2), encoding="utf-8")

    init_db()
    save_model_snapshot(snapshot)
    replace_sheet_rows("Dimensions", snapshot["tables"]["dimensions"], today)
    replace_sheet_rows("Indicators", snapshot["tables"]["indicators"], today)
    replace_sheet_rows("Inputs", snapshot["tables"]["inputs"], today)
    replace_sheet_rows("Scores", snapshot["tables"]["scores"], today)
    replace_sheet_rows("Alerts", snapshot["tables"]["alerts"], today)
    replace_sheet_rows(
        "DailyReports",
        [
            {
                "Date": today,
                "AsOf": today,
                "TotalScore": total_score,
                "Status": model_status,
                "Summary": short_summary,
                "GeneratedAt": generated_at,
                "ReportPath": f"reports/{today}.html",
            }
        ],
        today,
    )

    if report_enabled:
        upsert_daily_report(
            report_date=today,
            text=report_text,
            meta={"score": total_score, "status": model_status, "summary": ai_short_zh or short_summary},
            report_path=f"reports/{today}.html",
            payload=today_entry.get("reportPayload"),
            ai_analysis={
                "short_summary": ai_short_zh or short_summary,
                "detailed_interpretation": ai_detailed_zh,
                "short_summary_zh": ai_short_zh,
                "short_summary_en": ai_short_en,
                "detailed_interpretation_zh": ai_detailed_zh,
                "detailed_interpretation_en": ai_detailed_en,
                "model": ai_model,
                "status": "ok" if not ai_error else "fallback",
                "generated_at": generated_at,
            },
        )
        upsert_daily_report_ai_insight(
            report_date=today,
            short_summary=ai_short_zh or short_summary,
            detailed_text=ai_detailed_zh,
            insight=ai_payload,
            status="ok" if not ai_error else "fallback",
            model=ai_model,
            prompt_version=AI_PROMPT_VERSION,
            generated_at=generated_at,
            error=ai_error,
        )
        # Track estimated token usage for the daily generation pipeline.
        # Approximation rule: 1 token ~= 4 characters.
        token_input_text = json.dumps(snapshot, ensure_ascii=False)
        token_output_text = report_text or ""
        est_input_tokens = max(1, ceil(len(token_input_text) / 4))
        est_output_tokens = max(1, ceil(len(token_output_text) / 4))
        log_token_usage(
            source="daily_refresh",
            model="internal-estimation",
            input_tokens=est_input_tokens,
            output_tokens=est_output_tokens,
            total_tokens=est_input_tokens + est_output_tokens,
            meta={
                "asOf": today,
                "generatedAt": generated_at,
                "mode": mode,
                "type": "estimated",
            },
            logged_at=generated_at,
        )

    print(f"DONE mode={mode} as_of={today} updated={len(updated)} failed={len(failed)} total_score={total_score}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Refresh macro data and generate daily report")
    parser.add_argument("--mode", choices=["full", "fetch-only", "report-only"], default="full")
    parser.add_argument("--date", default="", help="Date in YYYY-MM-DD")
    parser.add_argument("--strict-freshness", action="store_true", help="Fail run if online-fetchable indicators are not freshly verified.")
    parser.add_argument("--require-openai-ai", action="store_true", help="Fail run if AI analysis is not generated by OpenAI model.")
    args = parser.parse_args()
    run(
        mode=args.mode,
        report_date=(args.date or None),
        strict_freshness=bool(args.strict_freshness),
        require_openai_ai=bool(args.require_openai_ai),
    )
